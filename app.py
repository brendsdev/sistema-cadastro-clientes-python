from ctypes import cdll
from os import name
from turtle import clear

import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook


# Aparencia padrão do sistema

ctk.set_appearance_mode("dark") # define o modo de aparência padrão como "dark" (escuro).
ctk.set_default_color_theme("blue") # define o tema de cores padrão como "blue" (azul). 

class App(ctk.CTk): # Esta classe representa a janela principal da aplicação e contém os métodos para configurar a interface e os widgets do sistema.
    def __init__(self):
        super().__init__()

        self.layout_config() # Chama o método layout_config para configurar o layout da janela, como título e tamanho.
        self.apperance() # Chama o método apperance para configurar os widgets relacionados ao tema do sistema, como o label e o menu suspenso para selecionar o tema.
        self.all_system() # Chama o método all_system, que provavelmente contém a lógica para configurar os demais widgets e funcionalidades do sistema.



    def layout_config(self):
        self.title("Sistema de Cadastro de Clientes") # Título da janela
        self.geometry("700x500") # Tamanho da janela

# Tema do sistema, com opções para o usuário escolher entre "light", "dark" e "system". O tema é aplicado usando a função set_appearance_mode do CustomTkinter, que altera a aparência da interface de acordo com a escolha do usuário.

    def apperance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000' , '#fff']).place(x=50, y=430) # Label para o tema, com cor de fundo transparente e cor de texto personalizada)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)  # Menu suspenso para selecionar o tema, com opções "light", "dark" e "system". O comando change_apm é chamado quando uma opção é selecionada.   


    def all_system(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal").place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Gestão de Clientes", font=("Century Gothic bold", 24), text_color="#fff", fg_color="teal").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=80)


        ficheiro = pathlib.Path("clientes.xlsx") # Cria um objeto Path para o arquivo "clientes.xlsx" usando a biblioteca pathlib.

        if ficheiro.exists(): # Verifica se o arquivo "clientes.xlsx" existe usando o método exists() do objeto Path.
                pass # Se o arquivo existir, o código continua normalmente.

        else:
                ficheiro = Workbook() # Se o arquivo não existir, é criado um novo arquivo Excel usando a classe Workbook da biblioteca openpyxl.
                folha=ficheiro.active # O novo arquivo Excel é salvo com o nome "clientes.xlsx" usando o método save() da classe Workbook.
                folha['A1'] = "Nome Completo" # A célula A1 da folha ativa é preenchida com o texto "Nome Completo".
                folha['B1'] = "Contato" # A célula B1 da folha ativa é preenchida com o texto "Contato".
                folha['C1'] = "Idade" # A célula C1 da folha ativa é preenchida com o texto "Idade".
                folha['D1'] = "Gênero" # A célula D1 da folha ativa é preenchida com o texto "Gênero".
                folha['E1'] = "Endereço" # A célula E1 da folha ativa é preenchida com o texto "Endereço".
                folha['F1'] = "Observações" # A célula F1 da folha ativa é preenchida com o texto "Observações".
                ficheiro.save("clientes.xlsx") # O arquivo Excel é salvo com o nome "clientes.xlsx" usando o método save() da classe Workbook.

        def submit():
# Pegando dados dos entrys



            name =name_value.get()
            contact = contact_value.get()
            age = age_value.get()
            address = address_value.get()
            gender = gender_combobox.get()
            obs = obs_entry.get()

# Validando os campos de entrada para garantir que todos os campos sejam preenchidos antes de salvar os dados. Se algum campo estiver vazio, uma mensagem de erro é exibida.

            if name == "" or contact == "" or age == "" or address == "" or gender == "" or obs == "":
                messagebox.showerror("Erro", "Por favor, preencha todos os campos!") # Exibe uma mensagem de erro usando a função showerror da biblioteca messagebox, indicando que o usuário deve preencher todos os campos.
                return

# Validando os dados

            ficheiro = openpyxl.load_workbook("clientes.xlsx") # Essa função serve para abrir um arquivo Excel no Python.
            folha = ficheiro.active # Acessa a folha ativa do arquivo Excel
            row = folha.max_row + 1 # Calcula a próxima linha disponível na folha ativa, somando 1 ao número da última linha preenchida (max_row) para garantir que os novos dados sejam inseridos na linha correta.

            folha.cell(row=row, column=1, value=name) # Insere os dados coletados dos campos de entrada
            folha.cell(row=row, column=2, value=contact)
            folha.cell(row=row, column=3, value=age)
            folha.cell(row=row, column=4, value=gender)
            folha.cell(row=row, column=5, value=address)
            folha.cell(row=row, column=6, value=obs)
            ficheiro.save("clientes.xlsx") # Salva as alterações feitas no arquivo Excel "clientes.xlsx" usando o método save da biblioteca openpyxl.
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso!") # Exibe uma mensagem de sucesso usando a função showinfo da biblioteca messagebox, indicando que os dados foram salvos com sucesso.
        
            clear() # Chama a função clear para limpar os campos de entrada após o envio dos dados.

        def clear(): # Essa função é responsável por limpar os campos de entrada (entrys) e o combo box, definindo seus valores como vazios ou resetando-os para o estado inicial.
            name_value.set("")
            contact_value.set("")
            age_value.set("")
            address_value.set("")
            gender_combobox.set("")
            obs_entry.delete(0, END)


# Text variables
        name_value = StringVar()
        contact_value = StringVar()
        age_value = StringVar()
        address_value = StringVar()


# Entrys

        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contact_entry = ctk.CTkEntry(self, width=200,textvariable=contact_value, font=("Century Gothic bold", 16), fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width=150, textvariable=age_value, font=("Century Gothic bold", 16), fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width=200, textvariable=address_value, font=("Century Gothic bold", 16), fg_color="transparent")

# Combo Box
        gender_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino", "Outro"], width=150, font=("Century Gothic bold", 14)) # Combo box para selecionar o gênero, com opções "Masculino", "Feminino" e "Outro". O combo box tem uma largura de 150 pixels, fonte personalizada e cor de fundo transparente.
        gender_combobox.set("Gênero")  # Define o valor inicial do combo box como "Gênero"

# Entrada de observações ou placeholder

        obs_entry = ctk.CTkEntry(self, width=500, height=150, font=("Arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")


        entry_name = ctk.CTkEntry(self, width=200, placeholder_text="Nome Completo")
        entry_contact = ctk.CTkEntry(self, width=200, placeholder_text="Contato")
        entry_age = ctk.CTkEntry(self, width=200, placeholder_text="Idade")
        entry_gender = ctk.CTkEntry(self, width=200, placeholder_text="Gênero")
        entry_address = ctk.CTkEntry(self, width=200, placeholder_text="Endereço")
        entry_obs = ctk.CTkEntry(self, width=200, placeholder_text="Observações")
    

# Labels

        lb_name = ctk.CTkLabel(self, text="Nome Completo", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contact = ctk.CTkLabel(self, text="Contato", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_age = ctk.CTkLabel(self, text="Idade", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_gender = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_address= ctk.CTkLabel(self, text="Endereço", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs= ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])


        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420) 
        btn_clear = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)

# Posicionando os elementos na janela

        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contact.place(x=450, y=120)
        contact_entry.place(x=450, y=150)

        lb_age.place(x=300, y=190)
        age_entry.place(x=300, y=220)
        
        lb_gender.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_address.place(x=50, y=190)
        address_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=150, y=260)


    def change_apm(self, new_apearance_mode):
        ctk.set_appearance_mode(new_apearance_mode) # Altera o modo de aparência com base na escolha do usuário no menu suspenso.



if __name__ == "__main__":
    app = App()
    app.mainloop() # Inicia a aplicação, criando uma instância da classe App e chamando o método mainloop() para manter a janela aberta.

